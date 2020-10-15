from math import sqrt
import string
import openpyxl
from openpyxl import workbook
import warnings
import matplotlib.pyplot as plt
import numpy as np 


def rmse_metric(actual, predicted):
    sum_error = 0.0
    for i in range(len(actual)):
            prediction_error = predicted[i] - actual[i]
            sum_error += (prediction_error ** 2)
    mean_error = sum_error / float(len(actual))
    return sqrt(mean_error)
 
def evaluate_algorithm(dataset, algorithm):
    test_set = list()
    for row in dataset:
            row_copy = list(row)
            row_copy[-1] = None
            test_set.append(row_copy)
    #global predicted
    predicted = algorithm(dataset, test_set)
    #print(predicted)
    actual = [row[-1] for row in dataset]
    rmse = rmse_metric(actual, predicted)
    return predicted
 
def mean(values):
    return sum(values) / float(len(values))
 
def covariance(x, mean_x, y, mean_y):
    covar = 0.0
    for i in range(len(x)):
            covar += (x[i] - mean_x) * (y[i] - mean_y)
    return covar
 
def variance(values, mean):
    return sum([(x-mean)**2 for x in values])
 
def coefficients(dataset):
    x = [row[0] for row in dataset]
    y = [row[1] for row in dataset]
    x_mean, y_mean = mean(x), mean(y)
    b1 = covariance(x, x_mean, y, y_mean) / variance(x, x_mean)
    b0 = y_mean - b1 * x_mean
    return [b0, b1]
 
def simple_linear_regression(train, test):
    predictions = list()
    b0, b1 = coefficients(train)
    for row in test:
            yhat = b0 + b1 * row[0]
            predictions.append(yhat)
    return predictions
    
def max_v(sheetx):
    a=0
    for col in sheetx.iter_cols(min_row=2,min_col=3, max_col=1000000, max_row=2):
            for cell in col:
                if(cell.value != None):
                    a+=1
                else:
                    return a

if __name__ == "__main__":

    threshold_tf = float(80)

    warnings.filterwarnings("ignore")

    wb= openpyxl.load_workbook('node_database.xlsx')
    sheet = wb.get_sheet_by_name('Sheet1')
    no_of_records = max_v(sheet)
    x_axis=[]
    for i in range(no_of_records):
        x_axis.append(float(i+1))
    #print(no_of_records)

    #AVAILABILITY
    avail=[]
    #avail=[[0 for x in range(no_of_records)] for y in range(20)]
    for i in range(50):
        avail_temp=[]
        for col in sheet.iter_cols(min_row=i+2, min_col=3, max_row=i+2, max_col=no_of_records+2):
            for cell in col:
                avail_temp.append(float(cell.value)*20.0)
        avail.append(avail_temp)

            
    avail_dataset=[]
    #avail_dataset=[[0 for x in range(no_of_records)] for y in range(20)]
    for i in range(50):
        temp_dataset=[]
        avail_k=avail[i]
        for j in range(no_of_records):
            #print(j)
            #temp[0]=x_axis[j]
            kkk=avail_k[j]
            #print(temp)
            temp_dataset.append([float(j+1),kkk])
            #print(temp_dataset)


        avail_dataset.append(temp_dataset)

    #print(avail_dataset)
    #print(avail_dataset[0])
    avail_predicted = []
    new_avail=[]
    avail_slope = []
    for i in range(50):
        dataset=avail_dataset[i]
        #print(dataset)
        predicted_out = evaluate_algorithm(dataset, simple_linear_regression)
        avail_predicted.append(predicted_out)
        (x1,y1) = (1,predicted_out[1])
        (x2,y2) = (2,predicted_out[2])
        x_in = no_of_records+1 
        slope = (y2-y1)/(x2-x1)
        avail_slope.append(slope)
        y_out = slope*(x_in-x1)+y1
        new_avail.append(y_out)
        


    #INTEGRITY
    integ=[]
    #avail=[[0 for x in range(no_of_records)] for y in range(20)]
    for i in range(50):
        integ_temp=[]
        for col in sheet.iter_cols(min_row=i+55, min_col=3, max_row=i+55, max_col=no_of_records+2):
            for cell in col:
                integ_temp.append(float(cell.value)*0.2)
        integ.append(integ_temp)

    integ_dataset=[]
    #avail_dataset=[[0 for x in range(no_of_records)] for y in range(20)]
    for i in range(50):
        temp_dataset=[]
        integ_k=integ[i]
        for j in range(no_of_records):
            #print(j)
            #temp[0]=x_axis[j]
            kkk=integ_k[j]
            #print(temp)
            temp_dataset.append([float(j+1),kkk])
            #print(temp_dataset)


        integ_dataset.append(temp_dataset)

    new_integ=[]
    integ_predicted = []
    integ_slope = []
    for i in range(50):
        dataset=integ_dataset[i]
        #print(dataset)
        predicted_out = evaluate_algorithm(dataset, simple_linear_regression)
        integ_predicted.append(predicted_out)
        (x1,y1) = (1,predicted_out[1])
        (x2,y2) = (2,predicted_out[2])
        x_in = no_of_records+1 
        slope = (y2-y1)/(x2-x1)
        integ_slope.append(slope)
        y_out = slope*(x_in-x1)+y1
        new_integ.append(y_out)

    #SECURITY
    secur=[]
    
    #avail=[[0 for x in range(no_of_records)] for y in range(20)]
    for i in range(50):
        secur_temp=[]
        for col in sheet.iter_cols(min_row=i+108, min_col=3, max_row=i+108, max_col=no_of_records+2):
            for cell in col:
                secur_temp.append(float(cell.value)*20)
        secur.append(secur_temp)

    secur_dataset=[]

    for i in range(50):
        temp_dataset=[]
        secur_k=secur[i]
        for j in range(no_of_records):
            #print(j)
            #temp[0]=x_axis[j]
            kkk=secur_k[j]
            #print(temp)
            temp_dataset.append([float(j+1),kkk])
            #print(temp_dataset)
        secur_dataset.append(temp_dataset)
        
    new_secur=[]
    secur_predicted = []
    secur_slope = []
    for i in range(50):
        dataset=secur_dataset[i]
        #print(dataset)
        predicted_out = evaluate_algorithm(dataset, simple_linear_regression)
        secur_predicted.append(predicted_out)
        (x1,y1) = (1,predicted_out[1])
        (x2,y2) = (2,predicted_out[2])
        x_in = no_of_records+1 
        slope = (y2-y1)/(x2-x1)
        secur_slope.append(slope)
        y_out = slope*(x_in-x1)+y1
        new_secur.append(y_out)

    #HONESTY
    honest=[]
    #avail=[[0 for x in range(no_of_records)] for y in range(20)]
    for i in range(50):
        honest_temp=[]
        for col in sheet.iter_cols(min_row=i+161, min_col=3, max_row=i+161, max_col=no_of_records+2):
            for cell in col:
                honest_temp.append(float(cell.value)*0.2)
        honest.append(honest_temp)

    honest_dataset=[]
    #avail_dataset=[[0 for x in range(no_of_records)] for y in range(20)]
    for i in range(50):
        temp_dataset=[]
        honest_k=honest[i]
        for j in range(no_of_records):
            #print(j)
            #temp[0]=x_axis[j]
            kkk=honest_k[j]
            #print(temp)
            temp_dataset.append([float(j+1),kkk])
            #print(temp_dataset)


        honest_dataset.append(temp_dataset)

    new_honest=[]
    honest_predicted = []
    honest_slope = []
    for i in range(50):
        dataset=honest_dataset[i]
        #print(dataset)
        predicted_out = evaluate_algorithm(dataset, simple_linear_regression)
        honest_predicted.append(predicted_out)
        (x1,y1) = (1,predicted_out[1])
        (x2,y2) = (2,predicted_out[2])
        x_in = no_of_records+1 
        slope = (y2-y1)/(x2-x1)
        honest_slope.append(slope)
        y_out = slope*(x_in-x1)+y1
        new_honest.append(y_out)


    #PRIVACY
    priva=[]
    
    #avail=[[0 for x in range(no_of_records)] for y in range(20)]
    for i in range(50):
        priva_temp=[]
        for col in sheet.iter_cols(min_row=i+214, min_col=3, max_row=i+214, max_col=no_of_records+2):
            for cell in col:
                priva_temp.append(float(cell.value)*20)
        priva.append(priva_temp)

    priva_dataset=[]
   
    for i in range(50):
        temp_dataset=[]
        priva_k=priva[i]
        for j in range(no_of_records):
            #print(j)
            #temp[0]=x_axis[j]
            kkk=priva_k[j]
            #print(temp)
            temp_dataset.append([float(j+1),kkk])
            #print(temp_dataset)


        priva_dataset.append(temp_dataset)

    new_priva=[]
    priva_predicted = []
    priva_slope = []
    for i in range(50):
        dataset=priva_dataset[i]
        #print(dataset)
        predicted_out = evaluate_algorithm(dataset, simple_linear_regression)
        priva_predicted.append(predicted_out)
        (x1,y1) = (1,predicted_out[1])
        (x2,y2) = (2,predicted_out[2])
        x_in = no_of_records+1 
        slope = (y2-y1)/(x2-x1)
        priva_slope.append(slope)
        y_out = slope*(x_in-x1)+y1
        new_priva.append(y_out)

    


    tf=[]
    for i in range(50):
        tf.append(new_avail[i] + new_integ[i] + new_secur[i] + new_honest[i] + new_priva[i])

    
    print("Predicted Trust factor values:\n")
    for i in range(50):
        print("Node ID: \t" ,i+1)
        print("Availability:\t", (new_avail[i])/20.00)
        print("Integrity: \t", (new_integ[i])/0.2)
        print("Security: \t", (new_secur[i])/20.00)
        print("Honesty: \t", (new_honest[i])/0.2)
        print("Privacy: \t", (new_priva[i])/20)
        print("Trust value: \t", tf[i])
        print(" ")
    
    test_avail = avail[5]
    #print(test_avail)
    test_integ = integ[5]
    #print(test_integ)
    test_secur = secur[5]
    #print(test_secur)
    test_honest = honest[5]
    test_priva = priva[5]
    test_tf=[]

    for i in range(no_of_records):
        test_tf.append(test_avail[i]+test_integ[i]+test_secur[i] + test_honest[i] + test_priva[i])

    #print(test_tf)
    test_dataset=[]
    for i in range(no_of_records):
        test_dataset.append([i+1,test_tf[i]])
    


    dataset = test_dataset
    predicted_out = evaluate_algorithm(dataset, simple_linear_regression)
    
    plt.figure(1)
    plt.plot(x_axis,predicted_out)
    plt.plot(x_axis,test_tf,'ro')
    plt.axis([0,no_of_records,0,150])
    plt.xlabel('Record no.')
    plt.ylabel('Trust Factor')
    plt.title('Model for Linear regression to predict Trust Factor') 
    plt.show()

    avail_av = 0
    integ_av = 0
    secur_av = 0
    honest_av = 0
    priva_av = 0
    for i in range(10):
        avail_av+=(avail[i][5]*5)
        integ_av+=(integ[i][5]*5)
        secur_av+=(secur[i][5]*5)
        honest_av+=(honest[i][5]*5)
        priva_av+=(priva[i][5]*5)
    avail_av/=10
    integ_av/=10
    secur_av/=10
    honest_av/=10
    priva_av/=10
    tf_av = (avail_av + integ_av + secur_av + honest_av + priva_av)/5


    avail_av_20 = 0
    integ_av_20 = 0
    secur_av_20 = 0
    honest_av_20 = 0
    priva_av_20 = 0
    for i in range(20):
        avail_av_20+=(avail[i][5]*5)
        integ_av_20+=(integ[i][5]*5)
        secur_av_20+=(secur[i][5]*5)
        honest_av_20+=(honest[i][5]*5)
        priva_av_20+=(priva[i][5]*5)
    avail_av_20/=20
    integ_av_20/=20
    secur_av_20/=20
    honest_av_20/=20
    priva_av_20/=20
    tf_av_20 = (avail_av_20 + integ_av_20 + secur_av_20 + honest_av_20 + priva_av_20)/5
    

    avail_av_30 = 0
    integ_av_30 = 0
    secur_av_30 = 0
    honest_av_30 = 0
    priva_av_30 = 0
    for i in range(30):
        avail_av_30+=(avail[i][5]*5)
        integ_av_30+=(integ[i][5]*5)
        secur_av_30+=(secur[i][5]*5)
        honest_av_30+=(honest[i][5]*5)
        priva_av_30+=(priva[i][5]*5)
    avail_av_30/=30
    integ_av_30/=30
    secur_av_30/=30
    honest_av_30/=30
    priva_av_30/=30
    tf_av_30 = (avail_av_30 + integ_av_30 + secur_av_30 + honest_av_30 + priva_av_30)/5



    avail_av_40 = 0
    integ_av_40 = 0
    secur_av_40 = 0
    honest_av_40 = 0
    priva_av_40 = 0
    for i in range(40):
        avail_av_40+=(avail[i][5]*5)
        integ_av_40+=(integ[i][5]*5)
        secur_av_40+=(secur[i][5]*5)
        honest_av_40+=(honest[i][5]*5)
        priva_av_40+=(priva[i][5]*5)
    avail_av_40/=40
    integ_av_40/=40
    secur_av_40/=40
    honest_av_40/=40
    priva_av_40/=40
    tf_av_40 = (avail_av_40 + integ_av_40 + secur_av_40 + honest_av_40 + priva_av_40)/5


    avail_av_50 = 0
    integ_av_50 = 0
    secur_av_50 = 0
    honest_av_50 = 0
    priva_av_50 = 0
    for i in range(50):
        avail_av_50+=(avail[i][5]*5)
        integ_av_50+=(integ[i][5]*5)
        secur_av_50+=(secur[i][5]*5)
        honest_av_50+=(honest[i][5]*5)
        priva_av_50+=(priva[i][5]*5)
    avail_av_50/=50
    integ_av_50/=50
    secur_av_50/=50
    honest_av_50/=50
    priva_av_50/=50
    tf_av_50 = (avail_av_50 + integ_av_50 + secur_av_50 + honest_av_50 + priva_av_50)/5
    
    x_all = [10,20,30,40,50]
    avail_all = [avail_av,avail_av_20,avail_av_30,avail_av_40,avail_av_50]
    integ_all = [integ_av,integ_av_20,integ_av_30,integ_av_40,integ_av_50]
    secur_all = [secur_av,secur_av_20,secur_av_30,secur_av_40,secur_av_50]
    honest_all = [honest_av,honest_av_20,honest_av_30,honest_av_40,honest_av_50]
    priva_all = [priva_av,priva_av_20,priva_av_30,priva_av_40,priva_av_50]
    tf_all = [tf_av,tf_av_20,tf_av_30,tf_av_40,tf_av_50]

    #print(tf_all.sort())
    
    plt.figure(2)
    plt.plot(x_all,avail_all, 'b',label='Availablility')
    plt.plot(x_all,integ_all, 'g', label='Integrity')
    plt.plot(x_all,secur_all, 'r', label='Security')
    plt.plot(x_all,honest_all, 'm', label='Honesty')
    plt.plot(x_all,priva_all, 'y', label='Privacy')
    plt.plot(x_all,sorted(tf_all), 'k', label='Resultant TF')
    plt.axis([0,50,0,100])
    plt.xlabel('No. Of Nodes')
    plt.ylabel('Trust value (Scaled to 100)')
    plt.title("Trust values' trends over no. of nodes")
    plt.legend()
    plt.show()
    

    mal_wb= openpyxl.load_workbook('malicious_node_data.xlsx')
    mal_sheet = mal_wb.get_sheet_by_name('mal')
    for i in range(50):
        mal_sheet.cell(row=i+3,column=2).value=None
        mal_sheet.cell(row=i+3,column=4).value=None
        mal_sheet.cell(row=i+3,column=6).value=None
    
    mal_wb.save('malicious_node_data.xlsx')
    working_nodes = []
    malicious_nodes = []
    pot_mal_nodes = []

    for i in range(50):
        if(tf[i]>threshold_tf):
            working_nodes.append(i+1)
            if((avail_slope[i]+integ_slope[i]+secur_slope[i]+honest_slope[i]+priva_slope[i])<0):
                pot_mal_nodes.append(i+1)
        else:
            malicious_nodes.append(i+1)
    
    
    for i in range(len(working_nodes)):
        mal_sheet.cell(row=i+3,column=2).value = working_nodes[i]

    for i in range(len(malicious_nodes)):
        mal_sheet.cell(row=i+3,column=4).value = malicious_nodes[i]

    for i in range(len(pot_mal_nodes)):
        mal_sheet.cell(row=i+3,column=6).value = pot_mal_nodes[i]

    mal_wb.save('malicious_node_data.xlsx')
    
    
    max_tf=max(tf)
    selected_node=tf.index(max_tf)+1

    print("\n\nSelected node: NODE NO.", selected_node)
    print("Trust value: ", max_tf)

    print("\nENTER TRUST FACTOR PARAMETER VALUES AFTER USE OF THE SELECTED NODE:")
    actual_avail = float(input())
    actual_integ = float(input())
    actual_secur = float(input())
    actual_honest = float(input())
    actual_priva = float(input())
    

    print("\nUpdating database...........")

    for i in range(50):
        sheet.cell(row = i+2,column = no_of_records+3).value = sheet.cell(row = i+2, column = no_of_records+2).value
        sheet.cell(row = i+55,column = no_of_records+3).value = sheet.cell(row = i+55, column = no_of_records+2).value        
        sheet.cell(row = i+108,column = no_of_records+3).value = sheet.cell(row = i+108, column = no_of_records+2).value
        sheet.cell(row = i+161, column = no_of_records+3).value = sheet.cell(row = i+161, column = no_of_records+2).value
        sheet.cell(row = i+214, column = no_of_records+3).value = sheet.cell(row = i+214, column = no_of_records+2).value

    wb.save('node_database.xlsx')
    sheet.cell(row=selected_node+1, column = no_of_records+3).value = actual_avail
    sheet.cell(row=selected_node+54, column = no_of_records+3).value = actual_integ
    sheet.cell(row=selected_node+107, column = no_of_records+3).value = actual_secur
    sheet.cell(row=selected_node+160, column = no_of_records+3).value = actual_integ
    sheet.cell(row=selected_node+213, column = no_of_records+3).value = actual_secur

    wb.save('node_database.xlsx')
    print('\n\t- - - - DATABASE UPDATED - - - - ')
    n=input()










